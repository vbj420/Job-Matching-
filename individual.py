import openpyxl
#import nlp_check
import START
import nltk
from nltk import pos_tag, word_tokenize
from nltk.corpus import stopwords
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity



# Extracted skills and job description keyword tokens
def match(required,profile) :
# Convert skills and job description tokens to lowercase
    extracted_skills  = required 
    job_description_tokens = profile
    extracted_skills_lower = [skill.lower() for skill in extracted_skills]
    job_description_tokens_lower = [token.lower() for token in job_description_tokens]
    
    # Create a count vectorizer and transform the data
    vectorizer = CountVectorizer().fit_transform(extracted_skills_lower + job_description_tokens_lower)
    vectorized_skills = vectorizer.toarray()
    
    # Calculate cosine similarity between skills and job description tokens
    cosine_similarities = cosine_similarity(vectorized_skills)
    
    # Determine matches based on cosine similarity
    skill_matches = {}
    for i, skill in enumerate(extracted_skills_lower):
        similarity_scores = cosine_similarities[i][len(extracted_skills_lower):]
        matched_indices = [index for index, score in enumerate(similarity_scores) if score > 0]
        matched_tokens = [job_description_tokens_lower[index] for index in matched_indices]
        skill_matches[skill] = matched_tokens
    
    '''
    # Display skill matches
    for skill, matched_tokens in skill_matches.items():
        print(f"Skill: {skill}")
        print(f"Matched Tokens: {matched_tokens}")
        print()
    '''
    print(skill_matches)
    total_similarity = sum(cosine_similarities[:len(extracted_skills_lower), len(extracted_skills_lower):])
    average_similarity = total_similarity / len(extracted_skills_lower)
    # Set a threshold for what you consider a "good" match
    threshold = 0
   # print(total_similarity)
    # Print "yes" if the average similarity is above the threshold, else print "no"
    if (average_similarity > threshold).any():
        return "yes"
    else:
        print(average_similarity)
        return "no"
# Load the Excel workbook
workbook = openpyxl.load_workbook("job_data.xlsx")
sheet = workbook.active
# Tokenize and tag words
def extract_skills(job_description) :
    tokens = word_tokenize(job_description)
    tagged_tokens = pos_tag(tokens)    
    # Filter out stopwords and extract nouns
    stop_words = set(stopwords.words('english'))
    skills = set()
    for word, tag in tagged_tokens:
        if tag.startswith('NN') and word.lower() not in stop_words:
            skills.add(word.lower())
    return skills
# Display login prompt
print("Welcome to the Job Portal!")

# Display available jobs
print("\nAvailable Jobs: \n")
c=1
for row in sheet.iter_rows(min_row=2, values_only=True):
    
    print(c,")") 
    c+=1 
    job_title, requirements = row[2], row[4]
    print(f"Job Title: {job_title}")
    print("Requirements: ")
    print(requirements)
    print()

# Gather user's skills
print("\nEnter your skills (Enter 'END' on a separate line to finish):")
skills = []
while True:
    skill = input()
    if skill == "END":
        break
    skills.append(skill)

# Check eligibility for each job
flag = 0 
matchS  = 0
print("\nEligible Jobs:")
for row in sheet.iter_rows(min_row=2, values_only=True):
    job_title, requirements = row[2], row[4]
    required_skills = extract_skills(requirements)
#   print(required_skills)
    val  = match(required_skills,skills)
    #print(val)
    for i in skills  :
        if i not in required_skills :
                flag  +=1 
        if i in required_skills :
                matchS +=1
    if val=="yes"  :
        print(f"You are eligible for: {job_title}")
                
print("\nThank you for using the Job Portal!")
