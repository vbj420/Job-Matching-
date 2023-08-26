import openpyxl

# Load existing Excel workbook or create a new one if it doesn't exist
try:
    workbook = openpyxl.load_workbook("job_data.xlsx")
except FileNotFoundError:
    workbook = openpyxl.Workbook()

sheet = workbook.active

# Find the last row with data
last_row = sheet.max_row + 1

# Get input for job data
job_title = input("Job Title: ")
location = input("Location: ")
company = input("Company: ")

print("Enter Responsibilities (Enter 'END' on a separate line to finish):")
responsibilities_lines = []
while True:
    line = input()
    if line == "END":
        break
    responsibilities_lines.append(line)
responsibilities = "\n".join(responsibilities_lines)

print("Enter Requirements (Enter 'END' on a separate line to finish):")
requirements_lines = []
while True:
    line = input()
    if line == "END":
        break
    requirements_lines.append(line)
requirements = "\n".join(requirements_lines)

print("Enter Benefits (Enter 'END' on a separate line to finish):")
benefits_lines = []
while True:
    line = input()
    if line == "END":
        break
    benefits_lines.append(line)
benefits = "\n".join(benefits_lines)

# Store job data in separate columns
sheet.cell(row=last_row, column=1, value=job_title)
sheet.cell(row=last_row, column=2, value=location)
sheet.cell(row=last_row, column=3, value=company)
sheet.cell(row=last_row, column=4, value=responsibilities)
sheet.cell(row=last_row, column=5, value=requirements)
sheet.cell(row=last_row, column=6, value=benefits)

# Save the Excel file
workbook.save("job_data.xlsx")
print("Job data saved to job_data.xlsx")
